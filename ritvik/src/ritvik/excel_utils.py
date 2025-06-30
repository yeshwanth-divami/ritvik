from openpyxl import Workbook, load_workbook
from pathlib import Path
from typing import List, Optional
from .models import Task
from uuid import UUID
from datetime import datetime

EXCEL_FILE = Path(__file__).parent / "tasks.xlsx"

HEADERS = [
    "asigner", "asignee", "request_date", "task", "deadline", "status", "task_id"
]

def init_excel():
    if not EXCEL_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)

def add_task(task: Task):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        task.asigner,
        task.asignee,
        task.request_date.isoformat(),
        task.task,
        task.deadline.isoformat(),
        task.status,
        str(task.task_id)
    ])
    wb.save(EXCEL_FILE)

def list_tasks(asignee: Optional[str] = None) -> List[Task]:
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if asignee and row[1] != asignee:
            continue
        task = Task(
            asigner=row[0],
            asignee=row[1],
            request_date=datetime.fromisoformat(row[2]),
            task=row[3],
            deadline=datetime.fromisoformat(row[4]),
            status=row[5],
            task_id=UUID(row[6])
        )
        tasks.append(task)
    return tasks

def mark_task_complete(task_id: UUID, status: str = "complete") -> bool:
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if str(row[6].value) == str(task_id):
            row[5].value = status
            wb.save(EXCEL_FILE)
            return True
    return False

def update_task_status_by_name(asignee: str, task_name: str, status: str) -> bool:
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        # row[1] = asignee, row[3] = task
        if row[1].value == asignee and row[3].value == task_name:
            row[5].value = status
            wb.save(EXCEL_FILE)
            return True
    return False
